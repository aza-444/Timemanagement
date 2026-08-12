import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from typing import List, Tuple, Optional, Dict, Any
from database.models import Expense, Category


def generate_expenses_excel(expenses_data: List[Tuple[Expense, Optional[Category]]]) -> io.BytesIO:
    """
    Generates a beautifully styled Excel spreadsheet for user expenses and incomes.
    """
    records = []
    for exp, cat in expenses_data:
        tx_label = "🟢 Kirim" if exp.transaction_type == "income" else "🔴 Xarajat"
        records.append({
            "ID": exp.id,
            "Sana": exp.expense_date.strftime("%Y-%m-%d"),
            "Turi": tx_label,
            "Kategoriya": cat.name if cat else "Noma'lum",
            "Summa (so'm)": exp.amount,
            "Izoh": exp.description or "",
            "Kiritilgan vaqt": exp.created_at.strftime("%Y-%m-%d %H:%M") if exp.created_at else ""
        })

    df = pd.DataFrame(records)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Operatsiyalar', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Operatsiyalar']

        # Styling
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        total_font = Font(name="Calibri", size=11, bold=True, color="000000")
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        # Style headers
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                # Amount column formatting
                if col_num == 5:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num in [1, 2, 3, 7]:
                    cell.alignment = Alignment(horizontal="center")

        # Total summary row
        last_row = len(df) + 2
        worksheet.cell(row=last_row, column=4, value="JAMI:").font = total_font
        worksheet.cell(row=last_row, column=4).alignment = Alignment(horizontal="right")

        total_cell = worksheet.cell(row=last_row, column=5, value=f"=SUM(E2:E{last_row-1})")
        total_cell.font = total_font
        total_cell.number_format = '#,##0'
        total_cell.fill = total_fill
        total_cell.border = thin_border

        # Auto column width
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output.seek(0)
    return output


def create_sample_excel_template() -> io.BytesIO:
    """
    Creates a sample Excel file for users to fill in and import expenses/incomes.
    """
    sample_data = [
        {"Sana": "2026-08-11", "Turi": "🔴 Xarajat", "Kategoriya": "🍔 Oziq-ovqat", "Summa": 45000, "Izoh": "Tushlik"},
        {"Sana": "2026-08-11", "Turi": "🟢 Kirim",   "Kategoriya": "💡 Boshqa",     "Summa": 5000000, "Izoh": "Oylik Maosh"},
        {"Sana": "2026-08-10", "Turi": "🔴 Xarajat", "Kategoriya": "📚 Ta'lim",     "Summa": 120000, "Izoh": "Kitoblar xaridi"}
    ]
    df = pd.DataFrame(sample_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Import_Namuna', index=False)
        worksheet = writer.sheets['Import_Namuna']
        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = 20
    output.seek(0)
    return output


def parse_and_validate_excel(
    file_content: bytes,
    category_map: Dict[str, int]  # category_name (lowercase) -> category_id
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parses an uploaded Excel file, validates structure and rows.
    Returns (valid_records, error_messages)
    """
    errors = []
    valid_records = []

    try:
        df = pd.read_excel(io.BytesIO(file_content))
    except Exception as e:
        return [], [f"Excel faylini o'qib bo'lmadi. Fayl formati xato: {str(e)}"]

    if df.empty:
        return [], ["Excel fayli bo'sh!"]

    # Normalize column names
    col_mapping = {}
    for col in df.columns:
        c_clean = str(col).strip().lower()
        if "sana" in c_clean or "date" in c_clean:
            col_mapping[col] = "sana"
        elif "turi" in c_clean or "type" in c_clean or "tur" in c_clean or "operatsiya" in c_clean:
            col_mapping[col] = "turi"
        elif "kategoriya" in c_clean or "category" in c_clean or "turkum" in c_clean:
            col_mapping[col] = "kategoriya"
        elif "summa" in c_clean or "amount" in c_clean or "narx" in c_clean:
            col_mapping[col] = "summa"
        elif "izoh" in c_clean or "comment" in c_clean or "description" in c_clean or "eslatma" in c_clean:
            col_mapping[col] = "izoh"

    df = df.rename(columns=col_mapping)

    required_cols = ["sana", "kategoriya", "summa"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        return [], [f"Excel faylida majburiy ustunlar topilmadi: {', '.join(missing)}. Namuna faylidan foydalaning."]

    INCOME_KEYWORDS = ['oldim', 'topib olindi', 'qarzning qaytarilishi', 'avans', 'kirim', 'ish haqi', 'maosh', 'oluvdim', 'qaytardi']

    # Iterate over rows
    for index, row in df.iterrows():
        row_num = index + 2  # Excel row number 1-indexed header

        # 1. Validate Date
        raw_date = row.get("sana")
        exp_date = None
        if pd.isna(raw_date):
            errors.append(f"Satr {row_num}: Sana ko'rsatilmadi.")
            continue
        try:
            if isinstance(raw_date, (datetime, pd.Timestamp)):
                exp_date = raw_date.date()
            elif isinstance(raw_date, date):
                exp_date = raw_date
            else:
                parsed_dt = pd.to_datetime(str(raw_date).strip(), errors='coerce')
                if pd.isna(parsed_dt):
                    errors.append(f"Satr {row_num}: Sana formati xato ('{raw_date}'). YYYY-MM-DD shaklida bo'lishi kerak.")
                    continue
                exp_date = parsed_dt.date()
        except Exception:
            errors.append(f"Satr {row_num}: Sana o'qishda xatolik.")
            continue

        # 2. Validate Amount & Transaction Type
        raw_amount = row.get("summa")
        if pd.isna(raw_amount):
            errors.append(f"Satr {row_num}: Summa bo'sh.")
            continue
        try:
            amount = float(raw_amount)
        except ValueError:
            errors.append(f"Satr {row_num}: Summa son emas ('{raw_amount}').")
            continue

        # Determine type
        raw_type = str(row.get("turi", "")).strip().lower() if "turi" in df.columns else ""
        tx_type = "expense"

        if raw_type:
            if any(k in raw_type for k in ["kirim", "income", "+", "daromad", "gan"]):
                tx_type = "income"
            elif any(k in raw_type for k in ["xarajat", "expense", "-"]):
                tx_type = "expense"

        # If amount is negative, e.g. -50000 -> treated as expense or income
        if amount < 0:
            amount = abs(amount)

        # 3. Validate Category
        raw_cat = str(row.get("kategoriya", "")).strip()
        cat_id = None
        if raw_cat:
            raw_cat_lower = raw_cat.lower()
            # Direct or partial match
            for cat_name, cid in category_map.items():
                if cat_name.lower() in raw_cat_lower or raw_cat_lower in cat_name.lower():
                    cat_id = cid
                    break
        
        # If category not found, try default "boshqa" or first category
        if not cat_id and category_map:
            for cat_name, cid in category_map.items():
                if "boshqa" in cat_name.lower() or "other" in cat_name.lower():
                    cat_id = cid
                    break
            if not cat_id:
                cat_id = list(category_map.values())[0]

        # 4. Description
        raw_desc = row.get("izoh")
        desc = None if pd.isna(raw_desc) else str(raw_desc).strip()

        # Keyword-based type detection if 'turi' was not explicitly set in Excel
        if not raw_type and desc:
            desc_lower = desc.lower()
            if any(k in desc_lower for k in INCOME_KEYWORDS):
                if not (desc_lower.endswith('oldim') and any(word in desc_lower for word in ['shim', 'atir', 'taksi', 'ovqat', 'suv', 'puli', 'uchun'])):
                    tx_type = "income"

        valid_records.append({
            "expense_date": exp_date,
            "category_id": cat_id,
            "amount": amount,
            "description": desc,
            "transaction_type": tx_type
        })

    return valid_records, errors
